"""
Matching algorithm for grader preferences for the students of MS&E 271
Required - an XLSX sheet version of the google sheet of the google form that was filled by them

ALGORITHM:
- read in the CSV sheet
- go row by row and make assignments => this is where assignments are done by first come first served
  - you can ALSO use a lottery to randomly assign people a number
- Assign to the first person. If first person is full, then assign to the second, and so on in their choice list
  * if you go beyond their 5th choice, then randomly assign to the remaining teams which are NOT in their exclusion list

EDGE CASES:
- if someone enters multiple choices, take the one with the latest timestamp
- if someone has not entered all 5 grader choices or all 3 exclusion choices

TODO:
- Generate the sheet which has student names ---> grader assigned
- Create a verification script that checks this^ sheet for correctness

"""
import pandas, random, os
from openpyxl import Workbook

random.seed(42)

############## USER-CONTROLLED OPTIONS ###############
VERBOSE = False		# verbose print what's going on
RANDOMIZE_STUDENTS = True	# randomize the students before assigning (lottery system)

# DETAILS OF THE FORM
GRADER_CHOICE_LIST_FILE = "grader_choices_text.txt"
EXCLUDED_GRADERS_LIST_FILE = "excluded_grader_choices_text.txt"
GRADER_LIST_FILE = "graders.txt"

# master list of students
STUDENT_MASTER_LIST_FILE = "students_master_list_2018.txt"	# A comma separates pre-assigned graders if they exist
STUDENT_NAME_EMAIL_MAP_FILE = "student_name_email_mapping.csv"

CSV_NAME = "STLP_grader_preferences_GEMWIN18_Jan29_2225hrs.csv"		# change to whatever's appropriate
OUTPUT_FILENAME1 = 'GEMWIN18_grader_assignments.xlsx'
OUTPUT_FILENAME2 = 'GEMWIN18_student_assignments.xlsx'
######################################################

def log(message):
	# print this out only if VERBOSE == True
	if VERBOSE:
		print ' '.join([str(m) for m in message])

def assign(student_email, grader):
	# assign student to grader. 
	# student_info is the tuple (SUID, STUDENT_NAME)
	# grader is the grader team name - it should be one of the keys of the GRADER dict
	assert grader in GRADERS.keys()
	GRADERS[grader]["students"].append(student_email)
	STUDENT_GRADER[student_email] = grader
	log(["Assigned", student_email, "to grader", grader])

def read_from_file(filename, delimiter=None):
	# reads a file with a string on each line
	# returns as a list of strings
	with open(filename) as f:
		x = f.readlines()
	# remove newlines
	if not delimiter:
		return [y.strip() for y in x]
	return [y.strip().split(delimiter) for y in x]

def grader_has_spots_open(gname):
	# gname - string
	# can this grader take any more students?
	return len(GRADERS[gname]["students"]) < GRADERS[gname]["limit"]

GRADER_CHOICE_LIST = read_from_file(GRADER_CHOICE_LIST_FILE)
EXCLUDED_GRADERS_LIST = read_from_file(EXCLUDED_GRADERS_LIST_FILE)

SUID = "What is your SUNetID?"
STUDENT_MASTER_LIST = read_from_file(STUDENT_MASTER_LIST_FILE, delimiter=',')

# create a dictionary which will map student email address => grader
STUDENT_GRADER = {x:y if y != '' else None for (x,y) in  STUDENT_MASTER_LIST}
STUDENT_NAME_EMAIL = {email:name.replace('"', '') for (name, email) in read_from_file(STUDENT_NAME_EMAIL_MAP_FILE, delimiter='", ')}

DEFAULT_STUDENT_LIMIT = 17 # the default limit on the number of students for a particular grader team

# this is a list of the choices for graders, MUST BE FILLED BEFOREHAND
# mapping grader strings => {"students" : list of students, "limit": the upper limit of the number of students they can take}
GRADERS_MASTER_LIST = read_from_file(GRADER_LIST_FILE, delimiter=",")
GRADERS = {}

for g in GRADERS_MASTER_LIST:
	LIMIT = DEFAULT_STUDENT_LIMIT
	if len(g) > 1:
		# the second entry exists, which is the limit for this grader
		LIMIT = int(g[1])
	GRADERS[g[0]] = {"students" : [], "limit": LIMIT}

STUDENTS_WHO_GOT_THEIR_CHOICES = 0
STUDENTS_WHO_HAD_TO_BE_RANDOMLY_ASSIGNED = 0
TOTAL_STUDENTS = 0

# Now pre-assign the students who had been pre-selected
for student in STUDENT_GRADER:
	grader = STUDENT_GRADER[student]
	if grader is not None:
		assign(student, grader.strip())

# read in data
data = pandas.read_csv(CSV_NAME, parse_dates=["Timestamp"])

# only take the last entry by timestamp.
# group by SUID, take last
data = data.groupby(SUID).last().reset_index()

# OPTION 1: randomize 
# ref: https://stackoverflow.com/questions/29576430/shuffle-dataframe-rows
if RANDOMIZE_STUDENTS:
	data = data.sample(frac=1).reset_index(drop=True)

# OPTION 2: go in existing order (first-come first-served)
# in this case, comment out previous line

# PRE-ASSIGNMENTS
# In the case that some students have been preferentially
# allotted to certain grading teams, then they can be entered here
if __name__ == "__main__":
	for index, row in data.iterrows():
		TOTAL_STUDENTS += 1

		# clean the row[SUID]
		row[SUID] = row[SUID].strip().lower()

		# if the student has already been pre-assigned, skip
		if STUDENT_GRADER[row[SUID]] != None:
			STUDENTS_WHO_GOT_THEIR_CHOICES += 1
			continue
		
		choice_index = 0	# pointer for their choice index
		assignment_completed = False	# used to check if this student was assigned preferred grader. If not, 
		NUMBER_OF_CHOICES_SPECIFIED = sum([1 for j in range(len(GRADER_CHOICE_LIST)) if not pandas.isnull(row[GRADER_CHOICE_LIST[j]])])	# number of choices this student has specified, upto a max possible of 5

		# while you still have specified choices left, 
		# iterate over the choices specified
		# and see if you still have spots left
		while choice_index < NUMBER_OF_CHOICES_SPECIFIED:
			if pandas.isnull(row[GRADER_CHOICE_LIST[choice_index]]):
				# if this choice is null, we've reached the end of the choices. leave the loop.
				break

			# Check if the type of choice is a grader pair!
			if row[GRADER_CHOICE_LIST[choice_index]][:11] == "Grader Pair":
				# if so, assign to BOTH graders only if BOTH have spots open!
				# This is because the string is of the form --> 'Grader Pair 7: grader name1 and grader name2'
				grader_pair = row[GRADER_CHOICE_LIST[choice_index]].split(": ")[1].split(' and ')
				if grader_has_spots_open(grader_pair[0]) and grader_has_spots_open(grader_pair[1]):
					assign(row[SUID], grader_pair[0])
					assign(row[SUID], grader_pair[1])

					# if a grader pair has been assigned,
					# then we need to update this specifically for that student
					STUDENT_GRADER[row[SUID]] = row[GRADER_CHOICE_LIST[choice_index]]

					STUDENTS_WHO_GOT_THEIR_CHOICES += 1
					assignment_completed = True
					break
				else:
					choice_index += 1

			# keep incrementing till you have a grader with spots left
			elif grader_has_spots_open(row[GRADER_CHOICE_LIST[choice_index]]):
				# this grader HAS spots left. assign to this, and move onto next student.
				assign(row[SUID], row[GRADER_CHOICE_LIST[choice_index]])
				STUDENTS_WHO_GOT_THEIR_CHOICES += 1
				assignment_completed = True
				break
			else:
				# this grader goes NOT have spots left. Try next.
				choice_index += 1

		if not assignment_completed:
			# if you've reached here, that means none of the top 5 choices had any spots left
			# create a set of the remaining graders, excluding:
			# 1. the ones the student doesn't want
			# 2. the ones who don't have spots left
			# 
			# assign randomly to one of these.
			EXCLUDED_GRADERS = [row[EXCLUDED_GRADERS_LIST[j]] for j in range(len(EXCLUDED_GRADERS_LIST)) if not pandas.isnull(row[EXCLUDED_GRADERS_LIST[j]])] 
			EXCLUDED_GRADERS += [grader for grader in GRADERS.keys() if len(GRADERS[grader]["students"]) == GRADERS[grader]["limit"]]
			# log(["no of choices specified:", NUMBER_OF_CHOICES_SPECIFIED])
			# log(["no of exclusions specified:", len(EXCLUDED_GRADERS)])
			# print "excluded graders for", row[SUID], "--->", EXCLUDED_GRADERS

			REMAINING_GRADERS_LIST = set([grader for grader in GRADERS.keys() if grader not in EXCLUDED_GRADERS])

			# print "Choosing from", REMAINING_GRADERS_LIST, "for", row[SUID]

			# sort graders in ascending order of how many students they already have,
			# choose the first - this will ensure that no grader ends up with no assignments
			# RANDOMLY_CHOSEN_GRADER = random.sample(REMAINING_GRADERS_LIST, 1)[0]
			RANDOMLY_CHOSEN_GRADER = sorted(REMAINING_GRADERS_LIST, key = lambda x: len(GRADERS[x]["students"]))[0]
			assign(row[SUID], RANDOMLY_CHOSEN_GRADER)
			STUDENTS_WHO_HAD_TO_BE_RANDOMLY_ASSIGNED += 1

	print "COMPLETED ASSIGNMENTS! Here's the overview:"
	print "------------------------------------------------"
	print "Total students who made choices =", TOTAL_STUDENTS
	print "Students who got one of their choices =", STUDENTS_WHO_GOT_THEIR_CHOICES
	print "Students who did not get any of their choices =", STUDENTS_WHO_HAD_TO_BE_RANDOMLY_ASSIGNED
	print "------------------------------------------------"

	# Now will randomly assign the remaining students to graders
	# until all students in the MASTER LIST have been completed
	STUDENTS_WHO_HAD_NO_PREFERENCE = [student for student in STUDENT_GRADER if STUDENT_GRADER[student] == None]
	print "Now assigning the remaining", len(STUDENTS_WHO_HAD_NO_PREFERENCE), "students who didn't fill the form"

	for student in STUDENTS_WHO_HAD_NO_PREFERENCE:
		# make a list of graders based on who has spots left
		REMAINING_GRADERS_LIST = [grader for grader in GRADERS if len(GRADERS[grader]["students"]) < GRADERS[grader]["limit"]]

		# choose one grader from this list randomly
		RANDOMLY_CHOSEN_GRADER = random.sample(REMAINING_GRADERS_LIST, 1)[0]

		# assign this student to this grader, continue
		assign(student, RANDOMLY_CHOSEN_GRADER)

	SORTED_GRADERS_LIST = sorted(GRADERS.keys())

	print "NOTE: Graders who didn't get any STLPs assigned to them: ", [grader for grader in SORTED_GRADERS_LIST if len(GRADERS[grader]["students"]) == 0]
	print "TOTAL STLPs THAT HAVE TO BE CHECKED: ", sum([len(GRADERS[grader]["students"]) for grader in SORTED_GRADERS_LIST])
	print "GRADER ASSIGNMENT COUNTS:"
	for grader in SORTED_GRADERS_LIST:
		current_count = len(GRADERS[grader]["students"])
		print grader, ":", current_count

	# write out to workbook (FOR GRADERS)
	wb = Workbook()
	ws = wb.active
	wb.remove_sheet(ws)

	# print GRADERS

	for grader in SORTED_GRADERS_LIST:
		# create new sheet
		ws = wb.create_sheet(title=grader)
		ws.append((grader, ))
		ws.append(("STUDENT NAME",))

		for student in GRADERS[grader]["students"]:
			ws.append((STUDENT_NAME_EMAIL[student], ))

	wb.save(filename=os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_FILENAME1))
	print "Succesfully written grader assignments to", OUTPUT_FILENAME1

	# write out to workbook (FOR STUDENTS)
	wb = Workbook()
	ws = wb.active
	ws.append(("Student Name", "Email", "Assigned grader"))

	for student in sorted(STUDENT_GRADER):
		ws.append((STUDENT_NAME_EMAIL[student], "{}@stanford.edu".format(student), STUDENT_GRADER[student]))

	wb.save(filename=os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_FILENAME2))
	print "Succesfully written student <--> graders to", OUTPUT_FILENAME2
